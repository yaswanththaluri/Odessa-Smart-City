from datetime import date

from openpyxl import Workbook, load_workbook
import os.path




# --------------------------------Rover Selection --------------------------


class Rover:

    def __init__(self, i1, j1, power_left, distance_covered):
        self.i1 = i1
        self.j1 = j1
        self.power_left = power_left
        self.distance_covered = distance_covered

    def position_s(self, i, j):
        self.i1 = i
        self.j1 = j

    def position(self, i, j):
        self.i1 = i
        self.j1 = j

    def nearest(self, r, c):
        d = abs(self.i1 - r) + abs(c - self.j1)
        return d


class Shop:

    def __init__(self, i1, j1):
        self.i1 = i1
        self.j1 = j1


def roverSelection():
    rover = [Rover(30, 20, 3000, 0), Rover(50, 40, 3000, 0), Rover(60, 10, 3000, 0), Rover(80, 70, 3000, 0),
             Rover(90, 10, 3000, 0)]

    shops = [Shop(40, 10), Shop(50, 60), Shop(10, 50), Shop(10, 90), Shop(80, 10)]

    print("\n\nCoordinates of Rover Initially")

    for i in range(len(shops)):
        print(shops[i].i1, shops[i].j1)

    while True:
        r1 = int(input("\n\nEnter ur Location : row no"))
        c1 = int(input("Enter ur Location : col no"))

        r2 = int(input("\n\nEnter shop Location : row no"))
        c2 = int(input("Enter shop Location : col no"))

        flag = True

        for i in range(len(shops)):
            if shops[i].i1 == r2 and shops[i].j1 == c2:
                flag = False

        if flag:
            print("Shop is not availble")
            continue

        rov = 0

        min = 1000000
        x, y = 0, 0
        for i in range(5):
            d = rover[i].nearest(r2, c2)

            if d < min:
                min = d
                x = rover[i].i1
                y = rover[i].j1
                rov = i

        print("Min dis = " + str(min) + " pos " + str(x) + " " + str(y))

        costToShop = abs(x - r2) + abs(y - c2)
        costToCust = abs(r2 - r1) + abs(c2 - c1)

        costToCust *= 10
        costToShop *= 10


        print("\n\nCost to Customer " + str(costToCust))
        print("cost to shop " + str(costToShop))

        rover[rov].power_left -= costToShop

        if rover[rov].power_left < 2000:
            print("\nThere is a shortage of power...Charging Rover")
            rover[rov].power_left = 2000

        if rover[rov].power_left < (2000 + costToCust):
            print("\nPower of rover will become lower than base case")
            rover[rov].power_left = 2000 + costToCust

        rover[rov].power_left -= costToCust
        print("\nPower left for rover is " + str(rover[rov].power_left))
        rover[rov].i1 = r1
        rover[rov].j1 = c1

        rover[rov].distance_covered = costToCust

        return rover[rov].distance_covered







# ----------------------------Cost Calculation -----------------------


def cost_date(d):
    today = date.today()
    distance = d
    order_cost = 0
    d1 = today.strftime("%Y/%m/%d")
    if (distance <100):
        order_cost = distance * (0.05 / 100)
        print(order_cost, d1)
        return order_cost, d1
    elif 100 <= distance <=1000:
        order_cost = (distance//100) * (0.50)
        print(order_cost, d1)
        return order_cost, d1
    elif 1000 <= distance <=2000:
        rest_distance = distance-1000
        order_cost = (order_cost + (distance * 0.50) + rest_distance * 0.75 )/100
        print(order_cost, d1)
        return order_cost, d1
    else:
        rest_distance = distance-2000
        order_cost = (order_cost + (distance * 0.75) + rest_distance * 0.85)/100
        print(order_cost, d1)
        return order_cost, d1





# ------------------Income Report ----------------------------


def readDailyIncome():
    wb = load_workbook("IncomeFile.xlsx")
    dailyIncome = wb["DailyIncome"]

    maxCol = dailyIncome.max_row

    for i in range(2, maxCol + 1):
        date = dailyIncome.cell(row=i, column=1)
        income = dailyIncome.cell(row=i, column=2)
        print("\nDate -> "+str(date.value), end="")
        print("  Income -> " + str(income.value))


def addDailyIncome(datetoAdd, incomeObt):
    wb = load_workbook("IncomeFile.xlsx")
    dailyIncome = wb["DailyIncome"]

    maxRow = dailyIncome.max_row

    rowNo = -1
    for i in range(1, maxRow + 1):
        date = dailyIncome.cell(row=i, column=1)
        if date.value == datetoAdd:
            rowNo = i

    if rowNo != -1:
        initIncome = dailyIncome.cell(row=rowNo, column=2)
        dailyIncome.cell(row=rowNo, column=2, value=initIncome.value + incomeObt)
    else:
        dailyIncome.cell(row=maxRow + 1, column=1, value=datetoAdd)
        dailyIncome.cell(row=maxRow + 1, column=2, value=incomeObt)

    wb.save("IncomeFile.xlsx")

    addMonthlyIncome(datetoAdd[:-3], incomeObt)


def addMonthlyIncome(monthToAdd, incomeToAdd):
    wb = load_workbook("IncomeFile.xlsx")
    dailyIncome = wb["MonthlyIncome"]

    maxRow = dailyIncome.max_row

    rowNo = -1
    for i in range(1, maxRow + 1):
        date = dailyIncome.cell(row=i, column=1)
        if date.value == monthToAdd:
            rowNo = i

    if rowNo != -1:
        initIncome = dailyIncome.cell(row=rowNo, column=2)
        dailyIncome.cell(row=rowNo, column=2, value=initIncome.value + incomeToAdd)
    else:
        dailyIncome.cell(row=maxRow + 1, column=1, value=monthToAdd)
        dailyIncome.cell(row=maxRow + 1, column=2, value=incomeToAdd)

    wb.save("IncomeFile.xlsx")


def showMonthlyIncome():
    wb = load_workbook("IncomeFile.xlsx")
    dailyIncome = wb["MonthlyIncome"]

    maxCol = dailyIncome.max_row

    for i in range(2, maxCol + 1):
        date = dailyIncome.cell(row=i, column=1)
        income = dailyIncome.cell(row=i, column=2)
        print("\nMonth -> " + str(date.value), end="")
        print("  Income -> "+ str(income.value))


def showIncomeBetweenDates(startDate, endDate):
    wb = load_workbook("IncomeFile.xlsx")
    dailyIncome = wb["DailyIncome"]

    maxCol = dailyIncome.max_row

    temp = 0

    for i in range(2, maxCol + 1):
        date = dailyIncome.cell(row=i, column=1)
        income = dailyIncome.cell(row=i, column=2)
        if startDate <= date.value <= endDate:
            temp = 1
            print(date.value)
            print(income.value)

    if temp == 0:
        print("\n\nNo Data Found b/w these dates")











# ------------------------------ Main Code ----------------------------------





if not os.path.exists("IncomeFile.xlsx"):
    wb = Workbook()
    wb.save("IncomeFile.xlsx")
    dailyIncome = wb.create_sheet("DailyIncome", 0)
    monthlyIncome = wb.create_sheet("MonthlyIncome", 1)



print("\n\n\t\t-----------------ODESSA SMART CITY ------------------------")

while True:

    print("\n\n\n\n")
    print("Enter \n1. To see the daily Income\n2. Show Monthly Income\n3. Show Income Between Dates\n4. Delivery")
    n = int(input())

    if n == 1:
        readDailyIncome()
    elif n == 2:
        showMonthlyIncome()
    elif n == 3:
        startDate = input("Enter Start Date:\n")
        endDate = input("Enter End Date:\n")
        showIncomeBetweenDates(startDate, endDate)

    elif n == 4:
        di = roverSelection()
        res = cost_date(di)
        addDailyIncome(str(res[1]), res[0])

    else:
        break
